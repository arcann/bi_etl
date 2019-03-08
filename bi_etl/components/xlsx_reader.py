"""
Created on Apr 2, 2015
"""
import os

from bi_etl.scheduler.task import ETLTask
from openpyxl import load_workbook
from bi_etl.components.etlcomponent import ETLComponent
from datetime import datetime, time

__all__ = ['XLSXReader']


class XLSXReader(ETLComponent):
    """
    XLSXReader will read rows from an Microsoft Excel XLSX formatted sheet.
    
    Parameters
    ----------
    task: ETLTask
        The  instance to register in (if not None)
    
    file_name: str
        The file_name to parse as xlsx.
        
    logical_name: str
        The logical name of this source. Used for log messages.

    Attributes
    ----------
    column_names: list
        The names to use for columns
        
    header_row: int
        The sheet row to read headers from. Default = 1.
    
    start_row: int
        The first row to parse for data. Default = header_row + 1 
    
    workbook: :class:`openpyxl.workbook.workbook.Workbook`
        The workbook that was opened.
        
    log_first_row : boolean
        Should we log progress on the the first row read. *Only applies if Table is used as a source.*
        (inherited from ETLComponent)
        
    max_rows : int, optional
        The maximum number of rows to read. *Only applies if Table is used as a source.*
        (inherited from ETLComponent)
        
    primary_key: list
        The name of the primary key column(s). Only impacts trace messages.  Default=None.
        (inherited from ETLComponent)
    
    progress_frequency: int
        How often (in seconds) to output progress messages. None for no progress messages.
        (inherited from ETLComponent)
    
    progress_message: str
        The progress message to print. Default is ``"{logical_name} row # {row_number}"``.
        Note ``logical_name`` and ``row_number`` subs.
        (inherited from ETLComponent)
        
    restkey: str
        Column name to catch long rows (extra values).
        
    restval: str
        The value to put in columns that are in the column_names but 
        not present in a given row (missing values).     
    """
    def __init__(self,
                 task: ETLTask,
                 file_name: str,
                 logical_name: str=None,
                 **kwargs
                 ):
        self.file_name = file_name
        if logical_name is None:
            try: 
                logical_name = os.path.basename(self.file_name)
            except AttributeError:
                logical_name = str(self.file_name)
        
        # Don't pass kwargs up. They should be set here at the end
        super(XLSXReader, self).__init__(task=task,
                                         logical_name=logical_name,
                                        )

        self._column_names = None
        # column to catch long rows (more values than columns)
        self.restkey = 'extra data past last delimiter'
        # default value for short rows (value for missing keys)    
        self.restval = None    
                  
        self.__header_row = 1    
        self.__start_row = None
        self.__active_row = None               
        
        self.__workbook = None        
        self.__active_worksheet = None
        
        # Should be the last call of every init
        self.set_kwattrs(**kwargs)

    def __repr__(self):
        return "XLSXReader({})".format(self.logical_name)
    
    @property
    def header_row(self):
        """
        int
            The sheet row to read headers from. Default = 1.
        """
        return self.__header_row

    @header_row.setter
    def header_row(self, value):
        self.__header_row = value
        
    @property
    def start_row(self):
        """
        int
            The sheet row to start reading data from. Default = header_row + 1
        """
        if self.__start_row is not None:
            return self.__start_row
        else:
            return self.header_row + 1

    @start_row.setter
    def start_row(self, value):
        self.__start_row = value        

    @property
    def workbook(self):
        if self.__workbook is None:
            self.__workbook = load_workbook(filename=self.file_name, read_only=True)
        return self.__workbook
    
    def set_active_worksheet_by_name(self, sheet_name):
        self.__active_worksheet = self.workbook[sheet_name]
        self._column_names = None
        
    def set_active_worksheet_by_number(self, sheet_number):        
        sheet_name = self.get_sheet_names()[sheet_number]
        self.set_active_worksheet_by_name(sheet_name)
    
    @property
    def active_worksheet(self):
        if self.__active_worksheet is None:
            self.set_active_worksheet_by_number(0)
        return self.__active_worksheet            
    
    def get_sheet_names(self):
        return self.workbook.get_sheet_names()
    
    def get_sheet_by_name(self, name):
        """Returns a worksheet by its name.

        Parameters
        ----------
        name: str
            The name of the worksheet to look for
            
        Returns
        -------
        openpyxl.worksheet.worksheet.Worksheet
            Worksheet object, or None if no worksheet has the name specified.

        """
        try:
            return self.workbook[name]
        except KeyError:
            return
    
    @property
    def line_num(self):
        """
        The current line number in the source file.
        line_num differs from rows_read in that rows_read deals with rows that would be returned to the caller
        """
        return self.__active_row
    
    def _obtain_column_names(self):
        try:
            row = self.read_header_row()                
            self._column_names = row
            if self.trace_data:
                self.log.debug("Column names read: {}".format(self._column_names))
        except StopIteration:
            pass
    
    @staticmethod
    def _get_cell_value(cell):
        value = cell.value
        if hasattr(value, 'strip'):
            value = value.strip()
            if value == '':
                value = None
        elif isinstance(value, datetime):
            # Excel time values of 12:00:00 AM come in as 1899-12-30 instead
            if value == datetime(1899, 12, 30):
                value = time(12, 0, 0)
        return value
    
    @staticmethod
    def _get_cell_values(row_cells):
        # Convert empty strings to None to be consistent with DB reads
        return list(map(XLSXReader._get_cell_value, row_cells))
        
    def read_header_row(self):
        # See https://openpyxl.readthedocs.org/en/latest/tutorial.html
        row = next(self.active_worksheet.get_squared_range(1,
                                                           self.header_row,
                                                           None,
                                                           self.header_row,
                                                           )
                   )
        return XLSXReader._get_cell_values(row) 

    def _raw_rows(self):
        # See https://openpyxl.readthedocs.org/en/latest/tutorial.html
        self.__active_row = self.start_row
        this_iteration_header = self.generate_iteration_header()
        for row in self.active_worksheet.iter_rows(row_offset=self.start_row-1):
            self.__active_row += 1
            row_values = XLSXReader._get_cell_values(row)           
            d = self.Row(list(zip(self.column_names, row_values)), iteration_header=this_iteration_header)
            len_column_names = len(self.column_names)
            len_row = len(row_values)
            if len_column_names < len_row:
                if self.restkey is not None:
                    d[self.restkey] = row_values[len_column_names:]
            elif len_column_names > len_row:
                for key in self.column_names[len_row]:
                    d[key] = self.restval
            yield d 
            
    def close(self):
        super(XLSXReader, self).close()
