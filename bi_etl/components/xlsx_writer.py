"""
Created on Apr 2, 2015
"""
import re
import sys
import typing

from openpyxl import load_workbook, Workbook

from bi_etl.components.row.row import Row
from bi_etl.components.xlsx_reader import XLSXReader
from bi_etl.scheduler.task import ETLTask
from bi_etl.statistics import Statistics

__all__ = ['XLSXWriter']


class XLSXWriter(XLSXReader):
    """
    XLSXReader will read rows from an Microsoft Excel XLSX formatted sheet.
    
    Parameters
    ----------
    task: ETLTask
        The  instance to register in (if not None)
    
    file_name: str
        The file_name to parse as xlsx.
        
    logical_name: str
        The logical name of this source. Used for log messages.

    Attributes
    ----------
    column_names: list
        The names to use for columns
        
    header_row: int
        The sheet row to read headers from. Default = 1.
    
    start_row: int
        The first row to parse for data. Default = header_row + 1 
    
    workbook: :class:`openpyxl.workbook.workbook.Workbook`
        The workbook that was opened.
        
    log_first_row : boolean
        Should we log progress on the the first row read. *Only applies if Table is used as a source.*
        (inherited from ETLComponent)
        
    max_rows : int, optional
        The maximum number of rows to read. *Only applies if Table is used as a source.*
        (inherited from ETLComponent)
        
    primary_key: list
        The name of the primary key column(s). Only impacts trace messages.  Default=None.
        (inherited from ETLComponent)
    
    progress_frequency: int
        How often (in seconds) to output progress messages. None for no progress messages.
        (inherited from ETLComponent)
    
    progress_message: str
        The progress message to print. Default is ``"{logical_name} row # {row_number}"``.
        Note ``logical_name`` and ``row_number`` subs.
        (inherited from ETLComponent)
        
    restkey: str
        Column name to catch long rows (extra values).
        
    restval: str
        The value to put in columns that are in the column_names but 
        not present in a given row (missing values).     
    """
    def __init__(self,
                 task: typing.Optional[ETLTask],
                 file_name: str,
                 logical_name: str = None,
                 write_only: bool = True,
                 **kwargs
                 ):
        super().__init__(
            task=task,
            file_name=file_name,
            logical_name=logical_name,
        )
        self.headers_written = False
        self.write_only = write_only
        self._insert_cnt = 0
        self._insert_cnt_this_sheet = 0

        # Matches invalid XML1.0 unicode characters, like control characters:
        # http://www.w3.org/TR/2006/REC-xml-20060816/#charsets
        # http://stackoverflow.com/questions/1707890/fast-way-to-filter-illegal-xml-unicode-chars-in-python

        _illegal_unichrs = [(0x00, 0x08), (0x0B, 0x0C), (0x0E, 0x1F),
                            (0x7F, 0x84), (0x86, 0x9F),
                            (0xFDD0, 0xFDDF), (0xFFFE, 0xFFFF)]
        if sys.maxunicode >= 0x10000:  # not narrow build
            _illegal_unichrs.extend([(0x1FFFE, 0x1FFFF), (0x2FFFE, 0x2FFFF),
                                     (0x3FFFE, 0x3FFFF), (0x4FFFE, 0x4FFFF),
                                     (0x5FFFE, 0x5FFFF), (0x6FFFE, 0x6FFFF),
                                     (0x7FFFE, 0x7FFFF), (0x8FFFE, 0x8FFFF),
                                     (0x9FFFE, 0x9FFFF), (0xAFFFE, 0xAFFFF),
                                     (0xBFFFE, 0xBFFFF), (0xCFFFE, 0xCFFFF),
                                     (0xDFFFE, 0xDFFFF), (0xEFFFE, 0xEFFFF),
                                     (0xFFFFE, 0xFFFFF), (0x10FFFE, 0x10FFFF)])

        _illegal_ranges = ["%s-%s" % (chr(low), chr(high))
                           for (low, high) in _illegal_unichrs]
        self._illegal_xml_chars_RE = re.compile(u'[%s]' % u''.join(_illegal_ranges))

        # Should be the last call of every init
        self.set_kwattrs(**kwargs)

    def __repr__(self):
        return "XLSXWriter({})".format(self.logical_name)

    @property
    def workbook(self):
        if self._workbook is None:
            if self.write_only:
                self._workbook = Workbook(write_only=True)
            else:
                self._workbook = load_workbook(filename=self.file_name, read_only=False)
        return self._workbook

    def set_active_worksheet_by_name(self, sheet_name):
        if sheet_name not in self.workbook:
            self._active_worksheet = self.workbook.create_sheet(sheet_name)
            self._insert_cnt_this_sheet = 0
            if self.headers_written:
                self._column_names = None
                self.headers_written = False
        else:
            self._active_worksheet = self.workbook[sheet_name]
            self._column_names = None
            self._insert_cnt_this_sheet = 0

    def _obtain_column_names(self):
        raise ValueError(f'Column names must be explicitly set on {self}')

    def write_header(self):
        if self.column_names is None:
            raise ValueError("insert called before column_names set (or possibly column_names needs to be set after set_active_worksheet_by_name call.")
        self.active_worksheet.append(self.column_names)
        self.headers_written = True

    def insert_row(self,
                   source_row: Row,  # Must be a single row
                   stat_name: str = 'insert',
                   parent_stats: Statistics = None,
                   ):
        """
        Inserts a row into the database (batching rows as batch_size)

        Parameters
        ----------
        source_row
            The row with values to insert
        stat_name
        parent_stats

        Returns
        -------
        new_row
        """
        stats = self.get_stats_entry(stat_name, parent_stats=parent_stats)
        stats.timer.start()

        # row_values = [source_row[col] for col in self.column_names]

        if not self.headers_written:
            self.write_header()

        # Remove invalid XML characters
        values = []
        for column_name in self.column_names:
            try:
                col_value = source_row[column_name]
                if isinstance(col_value, str):
                    col_value = self._illegal_xml_chars_RE.sub('\\?', col_value)
            except KeyError:
                col_value = None
            values.append(col_value)

        self.active_worksheet.append(values)

        self._insert_cnt += 1
        self._insert_cnt_this_sheet += 1

        stats.timer.stop()

    def insert(self,
               source_row: typing.Union[Row, list],  # Could also be a whole list of rows
               parent_stats: Statistics = None,
               **kwargs
               ):
        """
        Insert a row or list of rows in the table.

        Parameters
        ----------
        source_row: :class:`Row` or list thereof
            Row(s) to insert
        parent_stats: bi_etl.statistics.Statistics
            Optional Statistics object to nest this steps statistics in.
            Default is to place statistics in the ETLTask level statistics.
        """

        if isinstance(source_row, list):
            for row in source_row:
                self.insert_row(
                    row,
                    parent_stats=parent_stats,
                    **kwargs
                )
        else:
            self.insert_row(
                 source_row,
                 parent_stats=parent_stats,
                 **kwargs
             )
            
    def close(self):
        if self.has_workbook_init():
            # add_table is not currently working.

            # table_range = CellRange(
            #     min_col=1, max_col=len(self.column_names),
            #     min_row=1, max_row=self._insert_cnt_this_sheet +1
            # )
            # table_name = f"{self.active_worksheet.title}_table".replace(' ', '_')
            # tab = Table(displayName=table_name, ref=table_range.coord)
            #
            # # Add a default style with striped rows and banded columns
            # style = TableStyleInfo(
            #     name='TableStyleMedium2',
            #     showRowStripes=True
            # )
            # tab.tableStyleInfo = style
            # self.active_worksheet.add_table(tab)

            self.workbook.save(filename=self.file_name)
            self.workbook.close()
        super().close()
